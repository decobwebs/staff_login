import os
import csv
import qrcode
from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook
import threading
from datetime import datetime, timedelta
from smtplib import SMTP
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__)

# Constants
LOGIN_CSV = "login_data.csv"
LOGOUT_CSV = "logout_data.csv"
EXCEL_FILE = "output/SessionDetails.xlsx"
QR_LOGIN_DIR = "qr_codes/login"
QR_LOGOUT_DIR = "qr_codes/logout"
ADMIN_EMAIL = "sadiqabuidris@gamail.com"  # Replace with actual admin email
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
MAIL_USERNAME = "team.afssltd@gmail.com"  # Replace with your email
MAIL_PASSWORD = "lajk rvva ytaa qsxe"  # Replace with your email password

# Ensure required directories exist
os.makedirs(QR_LOGIN_DIR, exist_ok=True)
os.makedirs(QR_LOGOUT_DIR, exist_ok=True)
os.makedirs("output", exist_ok=True)

# Function to create QR codes
def generate_qr_codes():
    for day in range(1, 31):
        date = (datetime.now() + timedelta(days=day)).strftime("%Y-%m-%d")
        login_qr = qrcode.make(f"Login QR for {date}")
        logout_qr = qrcode.make(f"Logout QR for {date}")

        login_path = os.path.join(QR_LOGIN_DIR, f"login_{date}.png")
        logout_path = os.path.join(QR_LOGOUT_DIR, f"logout_{date}.png")

        login_qr.save(login_path)
        logout_qr.save(logout_path)

        print(f"Generated QR codes for {date}: {login_path}, {logout_path}")

# Generate QR codes for the next 30 days
generate_qr_codes()

# Route for login
@app.route("/login", methods=["POST"])
def login():
    name = request.form["name"]
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(LOGIN_CSV, "a", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([name, timestamp])

    return jsonify({"message": "Login successful!"})

# Route for logout
@app.route("/logout", methods=["POST"])
def logout():
    name = request.form["name"]
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(LOGOUT_CSV, "a", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([name, timestamp])

    return jsonify({"message": "Logout successful!"})

# Convert CSV files to Excel
def convert_csv_to_excel(csv_files, excel_file):
    try:
        workbook = Workbook()
        for csv_file in csv_files:
            sheet_name = os.path.splitext(os.path.basename(csv_file))[0]
            sheet = workbook.create_sheet(title=sheet_name)

            with open(csv_file, "r") as f:
                reader = csv.reader(f)
                for row in reader:
                    sheet.append(row)

        # Remove default sheet if it's empty
        if "Sheet" in workbook.sheetnames and not workbook["Sheet"].max_row:
            del workbook["Sheet"]

        # Save Excel file
        workbook.save(excel_file)
        print(f"Excel file saved successfully: {excel_file}")
    except PermissionError as e:
        print(f"Permission error: {e}. Ensure the file is not open.")
    except Exception as e:
        print(f"Error during conversion: {e}")

# Send email to admin
def send_email_to_admin():
    try:
        convert_csv_to_excel([LOGIN_CSV, LOGOUT_CSV], EXCEL_FILE)

        msg = MIMEMultipart()
        msg["From"] = MAIL_USERNAME
        msg["To"] = ADMIN_EMAIL
        msg["Subject"] = "Session Details Report"

        body = "Attached are the session details (logins and logouts)."
        msg.attach(MIMEText(body, "plain"))

        with open(EXCEL_FILE, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(EXCEL_FILE)}")
            msg.attach(part)

        with SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(MAIL_USERNAME, MAIL_PASSWORD)
            server.send_message(msg)

        print("Email sent to admin successfully!")
    except Exception as e:
        print(f"Failed to send email: {e}")

# Monitor for inactivity (example implementation)
def monitor_inactivity():
    while True:
        try:
            # Simulating a daily task (e.g., send report at the end of the day)
            now = datetime.now()
            if now.hour == 23 and now.minute == 59:  # Example: Send email at 11:59 PM
                send_email_to_admin()
        except Exception as e:
            print(f"Inactivity monitoring error: {e}")

# Start the Flask application
if __name__ == "__main__":
    # Start inactivity monitoring in a separate thread
    threading.Thread(target=monitor_inactivity, daemon=True).start()

    # Run Flask app
    app.run(host="0.0.0.0", port=5000, debug=True)
